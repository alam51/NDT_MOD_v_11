from utils import LDD_CONNECTOR as CONNECTOR
import pandas as pd
from openpyxl import load_workbook


class EWIC:
    def __init__(self, start_date: str, end_date: str):
        self.CONNECTOR = CONNECTOR
        self.start_date = start_date
        self.end_date = end_date

        self.df_ampere = self._df_ampere()
        self.df_mw = self._df_mw()
        self.df_energy_reading = self._df_energy_reading()
        self.df_multiplication_factor = self.df_energy_reading.pivot(index='date', columns='name',
                                                                     values='multiplication_factor').iloc[-1, :]
        self.df_energy_reading_formatted = self._df_energy_reading_formatted()
        self.df_ampere_mw_fomatted = self._df_ampere_mw_fomatted()
        self.a = 4

    def _df_ampere(self) -> pd.DataFrame:
        ampere_query_str = f"""
            SELECT A.date_time, A.value AS Ampere, A_t.name
            FROM pgcbfinal.east_west_interconnector_amp_reading AS A 
            JOIN pgcbfinal.east_west_interconnector_amp_type AS A_t ON A.connector_id = A_t.id
            WHERE A.date_time BETWEEN '{self.start_date}' AND '{self.end_date}'
            """
        _df_ampere = pd.read_sql(ampere_query_str, CONNECTOR, index_col=['date_time'])
        return _df_ampere

    def _df_mw(self) -> pd.DataFrame:
        mw_query_str = f"""
            SELECT t.id, t.name, mw.value, mw.date_time FROM 
            pgcbfinal.east_west_interconnector_mw AS mw 
            JOIN pgcbfinal.east_west_interconnector_mw_mv_type AS t ON mw.connector_id = t.id
            WHERE mw.date_time BETWEEN '{self.start_date}' AND '{self.end_date}'
        """
        # df_mw = pd.read_sql(mw_query_str, CONNECTOR, index_col=['date_time', 'id'])
        _df_mw = pd.read_sql(mw_query_str, CONNECTOR, index_col=None)
        return _df_mw

    def _df_energy_reading(self) -> pd.DataFrame:
        _start_date = pd.to_datetime(self.start_date) - pd.to_timedelta('1d')
        energy_reading_query_str = f"""
                    SELECT l.id, l.name, r.date, r.value, l.multiplication_factor FROM
                    pgcbfinal.east_west_interconnector_line_reading AS r 
                    JOIN pgcbfinal.east_west_interconnector_line AS l ON r.line_id = l.id
                    WHERE r.date BETWEEN '{str(_start_date)}' AND '{self.end_date}'
                """
        df_energy_reading = pd.read_sql(energy_reading_query_str, CONNECTOR, index_col=None)
        return df_energy_reading

    def _df_energy_reading_formatted(self) -> pd.DataFrame:
        df_energy_reading_formatted_stage0 = self.df_energy_reading.pivot(index='date', columns='name', values='value')
        df_energy_reading_formatted_stage1 = df_energy_reading_formatted_stage0.fillna(method='bfill', axis=0)
        df_energy_reading_formatted = df_energy_reading_formatted_stage1.iloc[1:, :].copy(deep=True)
        for i, running_date in enumerate(df_energy_reading_formatted_stage1.index[1:]):
            previous_date = df_energy_reading_formatted_stage1.index[i]
            r_n_1 = df_energy_reading_formatted_stage1.loc[previous_date, :]
            r_n = df_energy_reading_formatted_stage1.loc[running_date, :]
            diff = r_n - r_n_1
            diff_mkwh = diff * self.df_multiplication_factor / 1e6
            df_energy_reading_formatted.loc[running_date, :] = diff_mkwh

        max_vals = df_energy_reading_formatted.max(axis=0)
        max_dates = df_energy_reading_formatted.idxmax(axis=0)
        min_vals = df_energy_reading_formatted.min(axis=0)
        min_dates = df_energy_reading_formatted.idxmin(axis=0)

        df_energy_reading_formatted.loc['max', :] = max_vals
        df_energy_reading_formatted.loc['max_dt', :] = max_dates
        df_energy_reading_formatted.loc['min', :] = min_vals
        df_energy_reading_formatted.loc['min_dt', :] = min_dates
        return df_energy_reading_formatted

    def _df_ampere_mw_fomatted(self) -> pd.DataFrame:
        df_mw_formatted = self.df_mw.pivot(index='date_time', columns='name', values='value')
        df_ampere_mw = self.df_ampere.join(df_mw_formatted, how='right').iloc[:, [0, 2, 3, 4]]  # actually 0:3
        # selected
        cols = df_ampere_mw.columns
        df_ampere_mw.columns = ['Ghorashal_' + cols[0], cols[1] + '_MW', cols[2] + '_MW', cols[3] + '_MW']
        for i in df_ampere_mw.index:
            if df_ampere_mw.loc[i, 'Ishurdi_MW'] > 0.0:
                df_ampere_mw.loc[i, 'Ghorashal_Ampere'] *= -1

        max_vals = df_ampere_mw.max(axis=0)
        max_dates = df_ampere_mw.idxmax(axis=0)
        min_vals = df_ampere_mw.min(axis=0)
        min_dates = df_ampere_mw.idxmin(axis=0)

        df_ampere_mw.loc['max_exp', :] = max_vals
        df_ampere_mw.loc['max_exp_dt', :] = max_dates
        df_ampere_mw.loc['max_imp', :] = min_vals * -1
        df_ampere_mw.loc['min_imp_dt', :] = min_dates
        return df_ampere_mw

    def write_to_excel(self, excel_path: str) -> None:
        with pd.ExcelWriter(excel_path) as workbook:
            self.df_ampere_mw_fomatted.to_excel(workbook, sheet_name='Amp_MW')
            self.df_energy_reading_formatted.to_excel(workbook, sheet_name='Energy_MkWh')

        workbook = load_workbook(excel_path)
        for sheet in workbook.sheetnames:
            # ws = workbook.sheetnames
            ws = workbook[sheet]
            freeze_cell = ws['B2']
            ws.freeze_panes = freeze_cell
        workbook.save(excel_path)
        workbook.close()


ewic = EWIC(start_date='2023-03-01', end_date='2023-03-31 23:00')
# ewic.write_to_excel(r'I:\My Drive\IMD\Monthly_Report\2023\2.February\EWIC\EWIC.xlsx')
ewic.write_to_excel(r'I:\My Drive\IMD\Monthly_Report\2023\3.March\EWIC\EWIC.xlsx')
# df_ampere = ewic.df_ampere()
# df_mw = ewic.df_mw()
# df_reading = ewic.df_energy_reading
# b = ewic.df_ampere_mw_fomatted
# c = ewic.df_energy_reading_formatted
a = 5
